
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO
import json
import os
import re
import shutil
from datetime import datetime

APP_VERSION = "Final Build - DS Loading Price Only"
DATA_DIR = "data"
RATE_FILE = os.path.join(DATA_DIR, "stock_rates.json")
BACKUP_DIR = os.path.join(DATA_DIR, "rate_backups")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

st.set_page_config(page_title="Excel Formula Fusion", layout="wide")

st.markdown("""
<style>
html, body, [class*="css"] { color:#111111 !important; }
.stApp { background:#f7f8fa; }
div[data-testid="stSidebar"] { background:#ffffff; }
button, .stDownloadButton button {
    background-color:#ff6b00 !important;
    color:white !important;
    border-radius:8px !important;
    font-weight:700 !important;
}
input, textarea, select {
    color:#111111 !important;
    background-color:#ffffff !important;
}
div[data-baseweb="select"] > div {
    color:#111111 !important;
    background-color:#ffffff !important;
}
label, p, span, div {
    color:#111111;
}
.small-note {
    padding: 0.75rem;
    border-radius: 8px;
    background: #fff3e6;
    border: 1px solid #ffb366;
}
</style>
""", unsafe_allow_html=True)

# -----------------------------
# Helpers
# -----------------------------
def load_rates():
    if not os.path.exists(RATE_FILE):
        return {}
    try:
        with open(RATE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def save_rates_with_backup(rates: dict):
    """Save rates and make a timestamped backup copy every time rates change."""
    old_rates = load_rates()
    changed = old_rates != rates

    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)

    # Backup previous current file before replacing it
    if os.path.exists(RATE_FILE):
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_DIR, f"stock_rates_{stamp}.json")
        shutil.copy2(RATE_FILE, backup_path)

    # Save latest file
    with open(RATE_FILE, "w", encoding="utf-8") as f:
        json.dump(rates, f, indent=2, ensure_ascii=False, sort_keys=True)

    # Also save a dated copy of the new version for audit trail
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    new_backup_path = os.path.join(BACKUP_DIR, f"stock_rates_saved_{stamp}.json")
    with open(new_backup_path, "w", encoding="utf-8") as f:
        json.dump(rates, f, indent=2, ensure_ascii=False, sort_keys=True)

    return changed, new_backup_path

def backup_list():
    if not os.path.exists(BACKUP_DIR):
        return []
    files = [f for f in os.listdir(BACKUP_DIR) if f.lower().endswith(".json")]
    files.sort(reverse=True)
    return files

def clean_text(v):
    if v is None:
        return ""
    return str(v).strip()

def safe_sheet_name(name):
    invalid = ['\\', '/', '*', '[', ']', ':', '?']
    for ch in invalid:
        name = name.replace(ch, "-")
    return name[:31]

def norm(s):
    return re.sub(r"\s+", " ", clean_text(s).upper())

def excel_quote_sheet(sheet):
    return "'" + sheet.replace("'", "''") + "'"

def detect_multiplier(text):
    """
    Returns (multiplier, status, reason)
    status:
      red = confident multiply
      orange = doubtful mention, do not multiply
      none = no multiplier
    """
    t = norm(text)
    if not t:
        return 1, "none", ""

    patterns = [
        (r"\bSET\s+OF\s+(\d+)\b", "set of"),
        (r"\b(\d+)\s*PACK\s*=\s*(\d+)\b", "pack equals"),
        (r"\b1\s*PACK\s*=\s*(\d+)\b", "1 pack equals"),
    ]

    m = re.search(r"\bSET\s+OF\s+(\d+)\b", t)
    if m:
        return int(m.group(1)), "red", f"Detected SET OF {m.group(1)}"

    m = re.search(r"\b(?:1\s*)?PACK\s*=\s*(\d+)\b", t)
    if m:
        return int(m.group(1)), "red", f"Detected PACK = {m.group(1)}"

    # Suspicious wording - flag orange but don't multiply
    if any(w in t for w in ["SET", "PACK", "BUNDLE", "KIT"]):
        return 1, "orange", "Possible set/pack wording, not multiplied"

    return 1, "none", ""

def workbook_sheet_names(uploaded_file):
    data = uploaded_file.getvalue()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=False)
    return wb.sheetnames

def load_workbooks(uploaded_file):
    data = uploaded_file.getvalue()
    wb_formula = load_workbook(BytesIO(data), data_only=False)
    wb_values = load_workbook(BytesIO(data), data_only=True)
    return wb_formula, wb_values

def make_stock_rates_sheet(wb, rates, mapping=None):
    name = "STOCK RATES"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    headers = ["Stock", "Rate", "Total SQM", "Total Price", "Last Generated"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFD966")
        cell.alignment = Alignment(horizontal="center")

    row = 2
    working_sheet_formula_name = None
    if mapping:
        working_sheet_formula_name = excel_quote_sheet(mapping["working_sheet"])
        start_col = mapping["start_col_letter"]
        end_col = mapping["end_col_letter"]
        stock_row = mapping["stock_row"]
        sqm_row = mapping["output_sqm_row"]
        price_row = mapping["output_price_row"]

    for stock, rate in sorted(rates.items()):
        ws.cell(row=row, column=1, value=stock)
        ws.cell(row=row, column=2, value=float(rate or 0))
        if mapping and working_sheet_formula_name:
            # Consolidated totals per stock.
            # These formulas update automatically if the stock rate changes in column B
            # or if any generated worksheet formula changes.
            ws.cell(row=row, column=3, value=f'=SUMIF({working_sheet_formula_name}!${start_col}${stock_row}:${end_col}${stock_row},A{row},{working_sheet_formula_name}!${start_col}${sqm_row}:${end_col}${sqm_row})')
            ws.cell(row=row, column=4, value=f'=SUMIF({working_sheet_formula_name}!${start_col}${stock_row}:${end_col}${stock_row},A{row},{working_sheet_formula_name}!${start_col}${price_row}:${end_col}${price_row})')
        else:
            ws.cell(row=row, column=3, value="")
            ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 1

    if row > 2:
        ref = f"A1:E{row-1}"
        tab = Table(displayName="StockRates", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    widths = {"A": 45, "B": 15, "C": 18, "D": 18, "E": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for r in range(2, row):
        ws.cell(row=r, column=2).number_format = '$#,##0.00'
        ws.cell(row=r, column=3).number_format = '0.00'
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
    return ws

def build_clean_qty_formula(col, qty_total_row, country_col, first_store_row, last_store_row, ignore_countries, multiplier=1):
    # total qty minus ignored-country rows only; blank country rows are not involved.
    terms = []
    for c in ignore_countries:
        c = c.upper().strip()
        if c:
            terms.append(f'SUMPRODUCT(({col}${first_store_row}:{col}${last_store_row})*(--(UPPER(${country_col}${first_store_row}:${country_col}${last_store_row})="{c}")))')
    subtract = "+".join(terms) if terms else "0"
    base = f"({col}${qty_total_row}-({subtract}))"
    if multiplier and multiplier != 1:
        return f"={base}*{multiplier}"
    return f"={base}"

def build_ds_formula(col, name_row, size_row, ref_sheet, ref_start_row, ref_end_row, ref_name_col, ref_size_col, ref_ds_col):
    rs = excel_quote_sheet(ref_sheet)
    return (
        f'=IFERROR(INDEX({rs}!${ref_ds_col}${ref_start_row}:${ref_ds_col}${ref_end_row},'
        f'MATCH(1,INDEX(({rs}!${ref_name_col}${ref_start_row}:${ref_name_col}${ref_end_row}={col}${name_row})'
        f'*({rs}!${ref_size_col}${ref_start_row}:${ref_size_col}${ref_end_row}={col}${size_row}),0),0)),"")'
    )

def build_stock_formula(col, name_row, size_row, ref_sheet, ref_start_row, ref_end_row, ref_name_col, ref_size_col, ref_stock_col):
    rs = excel_quote_sheet(ref_sheet)
    return (
        f'=IFERROR(INDEX({rs}!${ref_stock_col}${ref_start_row}:${ref_stock_col}${ref_end_row},'
        f'MATCH(1,INDEX(({rs}!${ref_name_col}${ref_start_row}:${ref_name_col}${ref_end_row}={col}${name_row})'
        f'*({rs}!${ref_size_col}${ref_start_row}:${ref_size_col}${ref_end_row}={col}${size_row}),0),0)),"")'
    )

def size_to_sqm_formula(size_cell_ref):
    # Excel formula: tries to extract W x H from strings like 1000 x 2200.
    # This is intentionally simple and editable in Excel.
    return (
        f'=IFERROR('
        f'VALUE(TRIM(LEFT(SUBSTITUTE(UPPER({size_cell_ref}),"X",REPT(" ",50)),50)))'
        f'*VALUE(TRIM(MID(SUBSTITUTE(UPPER({size_cell_ref}),"X",REPT(" ",50)),51,50)))'
        f'/1000000,"")'
    )

def build_price_formula(col, sqm_row, ds_row, stock_row, ds_loading_pct):
    factor = 1 + float(ds_loading_pct) / 100
    # Uses central STOCK RATES sheet. Change STOCK RATES once; all formulas update.
    rate_lookup = f'IFERROR(VLOOKUP({col}${stock_row},\'STOCK RATES\'!$A:$B,2,FALSE),0)'
    return f'=IF(UPPER({col}${ds_row})="DS",{col}${sqm_row}*{rate_lookup}*{factor},{col}${sqm_row}*{rate_lookup})'

def find_nonempty_cols(ws_values, row, start_col, end_col):
    cols = []
    for c in range(start_col, end_col + 1):
        if clean_text(ws_values.cell(row=row, column=c).value):
            cols.append(c)
    return cols


def parse_sum_range_from_total_formula(formula):
    """Return (start_row, end_row) from formulas like =SUM(AC8:AC153)."""
    if not formula or not isinstance(formula, str):
        return None
    m = re.search(r"SUM\(\$?[A-Z]+\$?(\d+)\s*:\s*\$?[A-Z]+\$?(\d+)\)", formula.upper())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))

def auto_detect_store_range(ws_formula, qty_total_row, start_col_idx, end_col_idx):
    """Detect store qty rows from the first total-row SUM formula in the item columns."""
    for c in range(start_col_idx, end_col_idx + 1):
        rng = parse_sum_range_from_total_formula(ws_formula.cell(row=qty_total_row, column=c).value)
        if rng:
            return rng
    return 8, max(8, qty_total_row + 146)

def validate_output_rows(mapping):
    rows = {
        "DS/SS output row": mapping["output_ds_row"],
        "Clean qty output row": mapping["output_clean_qty_row"],
        "SQM output row": mapping["output_sqm_row"],
        "Price output row": mapping["output_price_row"],
    }
    seen = {}
    for label, row in rows.items():
        if row in seen:
            raise ValueError(f"Output row conflict: {label} and {seen[row]} are both row {row}.")
        seen[row] = label
    protected = {
        mapping["name_row"]: "Name row",
        mapping["size_row"]: "Size row",
        mapping["stock_row"]: "Stock/material row",
        mapping["qty_total_row"]: "Original total qty row",
    }
    for label, row in rows.items():
        if row in protected:
            raise ValueError(f"Circular risk: {label} cannot overwrite {protected[row]} ({row}).")
        if mapping["first_store_row"] <= row <= mapping["last_store_row"]:
            raise ValueError(f"Circular risk: {label} row {row} is inside store qty rows {mapping['first_store_row']}:{mapping['last_store_row']}.")

# -----------------------------
# UI
# -----------------------------
st.title("Excel Formula Fusion")
st.caption(APP_VERSION)

st.markdown(
    '<div class="small-note"><b>Storage warning:</b> rates are saved to <code>data/stock_rates.json</code> and dated backups are saved in <code>data/rate_backups</code>. On Streamlit Cloud, redeploys or server resets may remove local files. Download backups regularly or move rates to Google Sheets/database later.</div>',
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("1. Upload")
    uploaded = st.file_uploader("Upload Excel workbook", type=["xlsx", "xlsm"])
    if uploaded:
        st.success(f"Uploaded: {uploaded.name}")

    st.header("Rate JSON")
    rate_upload = st.file_uploader("Upload stock_rates.json backup", type=["json"], key="rate_json_upload")
    if rate_upload and st.button("Restore uploaded rates"):
        try:
            restored = json.loads(rate_upload.getvalue().decode("utf-8"))
            if isinstance(restored, dict):
                changed, backup_path = save_rates_with_backup(restored)
                st.success("Rates restored and backup saved.")
                st.rerun()
            else:
                st.error("Uploaded JSON must be an object like {stock: rate}.")
        except Exception as e:
            st.error(f"Could not restore rates: {e}")

    current_rates = load_rates()
    st.download_button(
        "Download current stock_rates.json",
        data=json.dumps(current_rates, indent=2, ensure_ascii=False).encode("utf-8"),
        file_name="stock_rates.json",
        mime="application/json",
    )

if not uploaded:
    st.info("Upload an Excel workbook to begin.")
    st.stop()

# Read workbook sheets only
try:
    sheet_names = workbook_sheet_names(uploaded)
except Exception as e:
    st.error("Could not read workbook sheets.")
    st.exception(e)
    st.stop()

st.success(f"Workbook detected. Sheets: {len(sheet_names)}")

# Defaults
default_working = sheet_names.index("DL ANZ ALLOCATION") if "DL ANZ ALLOCATION" in sheet_names else 0
default_ref = sheet_names.index("PRINT DB") if "PRINT DB" in sheet_names else min(1, len(sheet_names)-1)

with st.form("mapping_form"):
    st.header("2. Mapping setup")

    c1, c2 = st.columns(2)
    with c1:
        working_sheet = st.selectbox("Working sheet", sheet_names, index=default_working)
        start_col_letter = st.text_input("Start Column", value="AC")
        end_col_letter = st.text_input("End Column", value="IG")
        name_row = st.number_input("Name row", min_value=1, value=4, step=1)
        size_row = st.number_input("Size row", min_value=1, value=5, step=1)
        stock_row = st.number_input("Stock/material row", min_value=1, value=6, step=1)
        qty_total_row = st.number_input("Original total qty row", min_value=1, value=7, step=1)
        country_col = st.text_input("Country column", value="I")
        first_store_row = st.number_input("First store row for ignored-country subtraction", min_value=1, value=8, step=1)
        last_store_row = st.number_input("Last store row for ignored-country subtraction", min_value=1, value=153, step=1)
        st.caption("Tip: for ASICS row 7 totals usually use SUM(row 8:153). Keep output rows outside this range.")

    with c2:
        ref_sheet = st.selectbox("Reference sheet", sheet_names, index=default_ref)
        ref_start_row = st.number_input("Reference start row", min_value=1, value=12, step=1)
        ref_end_row = st.number_input("Reference end row", min_value=1, value=141, step=1)
        ref_name_col = st.text_input("Reference name/artwork column", value="C")
        ref_size_col = st.text_input("Reference size column", value="E")
        ref_ds_col = st.text_input("Reference DS/SS column", value="F")
        ref_stock_col = st.text_input("Reference stock/material column", value="G")
        ignore_countries_text = st.text_input("Ignore countries, comma separated", value="NZ")
        ds_loading_pct = st.number_input("DS loading %", min_value=0.0, max_value=500.0, value=20.0, step=1.0)

    output_ds_row = st.number_input("Output DS/SS row", min_value=1, value=168, step=1)
    output_clean_qty_row = st.number_input("Output clean qty row", min_value=1, value=169, step=1)
    output_sqm_row = st.number_input("Output SQM row", min_value=1, value=170, step=1)
    output_price_row = st.number_input("Output price row", min_value=1, value=171, step=1)

    apply_mapping = st.form_submit_button("Apply Mapping / Load Stock List")

if "mapping" not in st.session_state or apply_mapping:
    try:
        st.session_state.mapping = {
            "working_sheet": working_sheet,
            "ref_sheet": ref_sheet,
            "start_col_letter": start_col_letter.upper().strip(),
            "end_col_letter": end_col_letter.upper().strip(),
            "name_row": int(name_row),
            "size_row": int(size_row),
            "stock_row": int(stock_row),
            "qty_total_row": int(qty_total_row),
            "country_col": country_col.upper().strip(),
            "first_store_row": int(first_store_row),
            "last_store_row": int(last_store_row),
            "ref_start_row": int(ref_start_row),
            "ref_end_row": int(ref_end_row),
            "ref_name_col": ref_name_col.upper().strip(),
            "ref_size_col": ref_size_col.upper().strip(),
            "ref_ds_col": ref_ds_col.upper().strip(),
            "ref_stock_col": ref_stock_col.upper().strip(),
            "ignore_countries": [x.strip().upper() for x in ignore_countries_text.split(",") if x.strip()],
            "ds_loading_pct": float(ds_loading_pct),
            "output_ds_row": int(output_ds_row),
            "output_clean_qty_row": int(output_clean_qty_row),
            "output_sqm_row": int(output_sqm_row),
            "output_price_row": int(output_price_row),
        }
        st.success("Mapping applied.")
    except Exception as e:
        st.error("Mapping error.")
        st.exception(e)
        st.stop()

m = st.session_state.mapping

# Load value workbook only when needed to find stock names
try:
    wb_formula_tmp, wb_values_tmp = load_workbooks(uploaded)
    ws_values = wb_values_tmp[m["working_sheet"]]
    start_idx = column_index_from_string(m["start_col_letter"])
    end_idx = column_index_from_string(m["end_col_letter"])
    item_cols = find_nonempty_cols(ws_values, m["name_row"], start_idx, end_idx)
    detected_range = auto_detect_store_range(wb_formula_tmp[m["working_sheet"]], m["qty_total_row"], start_idx, end_idx)

    stock_values = []
    for c in item_cols:
        v = clean_text(ws_values.cell(row=m["stock_row"], column=c).value)
        if v:
            stock_values.append(v)
    unique_stocks = sorted(set(stock_values))
    st.info(f"Detected store qty range from total row formula: rows {detected_range[0]}:{detected_range[1]}. If your mapping shows a different range, update it before export.")
    del wb_formula_tmp, wb_values_tmp
except Exception as e:
    st.error("Could not load stock list from selected mapping.")
    st.exception(e)
    st.stop()

st.header("3. Stock/material rates")
st.write("Stock names are scanned only from your selected Start Column to End Column.")

selected_stocks = st.multiselect(
    "Pick stock/material to calculate SQM and rate",
    unique_stocks,
    default=unique_stocks[: min(8, len(unique_stocks))]
)

current_rates = load_rates()
edited_rates = dict(current_rates)

if selected_stocks:
    st.subheader("Enter rates")
    with st.form("rate_form"):
        cols = st.columns(3)
        for i, stock in enumerate(selected_stocks):
            with cols[i % 3]:
                existing = float(current_rates.get(stock, 0) or 0)
                edited_rates[stock] = st.number_input(
                    f"{stock}",
                    min_value=0.0,
                    value=existing,
                    step=0.5,
                    format="%.2f",
                    key=f"rate_{stock}",
                )
        save_rates = st.form_submit_button("Save / Update Rates")
    if save_rates:
        cleaned_rates = {k: float(v or 0) for k, v in edited_rates.items() if clean_text(k)}
        changed, backup_path = save_rates_with_backup(cleaned_rates)
        st.success(f"Rates saved. Backup created: {backup_path}")
        st.rerun()
else:
    st.info("Pick one or more stocks to enter rates.")

with st.expander("Rate backups"):
    files = backup_list()
    if files:
        chosen_backup = st.selectbox("Available backups", files)
        path = os.path.join(BACKUP_DIR, chosen_backup)
        with open(path, "rb") as f:
            st.download_button("Download selected backup", f.read(), file_name=chosen_backup, mime="application/json")
    else:
        st.write("No backups yet.")

st.header("4. Generate workbook")

st.write("Current mapping:")
st.json(m)

if st.button("Generate Excel Workbook"):
    try:
        validate_output_rows(m)
        wb, wb_values = load_workbooks(uploaded)
        ws = wb[m["working_sheet"]]
        ws_values = wb_values[m["working_sheet"]]

        make_stock_rates_sheet(wb, load_rates(), m)

        start_idx = column_index_from_string(m["start_col_letter"])
        end_idx = column_index_from_string(m["end_col_letter"])

        red_fill = PatternFill("solid", fgColor="FF9999")
        orange_fill = PatternFill("solid", fgColor="F4B183")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        green_fill = PatternFill("solid", fgColor="C6EFCE")

        audit_name = "Qty Multiplier Audit"
        if audit_name in wb.sheetnames:
            del wb[audit_name]
        audit = wb.create_sheet(audit_name)
        audit.append(["Column", "Name", "Multiplier", "Status", "Reason"])
        for cell in audit[1]:
            cell.font = Font(bold=True)
            cell.fill = yellow_fill

        for c in range(start_idx, end_idx + 1):
            col = get_column_letter(c)
            name_value = ws_values.cell(row=m["name_row"], column=c).value
            if not clean_text(name_value):
                continue

            multiplier, status, reason = detect_multiplier(name_value)

            # DS/SS lookup formula
            ws.cell(row=m["output_ds_row"], column=c).value = build_ds_formula(
                col, m["name_row"], m["size_row"],
                m["ref_sheet"], m["ref_start_row"], m["ref_end_row"],
                m["ref_name_col"], m["ref_size_col"], m["ref_ds_col"]
            )

            # Keep original stock/material row untouched.
            # The working sheet already links row 6 to PRINT DB in current workbooks.
            # Price formulas reference this row, so changing STOCK RATES updates prices without rebuilding.

            # Clean qty
            ws.cell(row=m["output_clean_qty_row"], column=c).value = build_clean_qty_formula(
                col,
                m["qty_total_row"],
                m["country_col"],
                m["first_store_row"],
                m["last_store_row"],
                m["ignore_countries"],
                multiplier=multiplier,
            )

            # SQM = clean qty * size area ONLY. Do NOT apply DS loading here.
            # DS loading belongs only in the price formula below.
            area_formula = size_to_sqm_formula(f"{col}${m['size_row']}")
            # strip leading =
            ws.cell(row=m["output_sqm_row"], column=c).value = f"={col}${m['output_clean_qty_row']}*({area_formula[1:]})"

            # Price references central STOCK RATES sheet and applies DS loading only if DS row says DS.
            # Pattern: SQM * Rate * IF(DS, 1.2, 1)
            ws.cell(row=m["output_price_row"], column=c).value = build_price_formula(
                col, m["output_sqm_row"], m["output_ds_row"], m["stock_row"], m["ds_loading_pct"]
            )

            # Styles
            for rr in [m["output_ds_row"], m["output_clean_qty_row"], m["output_sqm_row"], m["output_price_row"]]:
                ws.cell(row=rr, column=c).fill = green_fill

            if status == "red":
                for rr in [m["name_row"], m["output_clean_qty_row"]]:
                    ws.cell(row=rr, column=c).fill = red_fill
                audit.append([col, clean_text(name_value), multiplier, "MULTIPLIED", reason])
            elif status == "orange":
                ws.cell(row=m["name_row"], column=c).fill = orange_fill
                audit.append([col, clean_text(name_value), multiplier, "CHECK - NOT MULTIPLIED", reason])

        # Add headings in output rows
        ws.cell(row=m["output_ds_row"], column=1).value = "DS/SS"
        ws.cell(row=m["output_clean_qty_row"], column=1).value = "Clean Qty"
        ws.cell(row=m["output_sqm_row"], column=1).value = "SQM"
        ws.cell(row=m["output_price_row"], column=1).value = "Price"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        st.success("Workbook generated.")
        st.download_button(
            "Download Excel Workbook",
            data=output.getvalue(),
            file_name="excel_formula_fusion_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        st.error("Generation failed.")
        st.exception(e)
